import openpyxl
from openpyxl.utils import get_column_letter

from collections import defaultdict

from matplotlib import pyplot as plot

from finance_tracker_project.config.config import REPORTS_DIR
from finance_tracker_project.errors.NoReportableTransactionsError import NoReportableTransactionsError
from finance_tracker_project.models.bank_account import BankAccount


class ReportsService:
    # rate on: 05.06.2025 15:41
    EXCHANGE_RATES_TO_PLN = {
        "PLN": 1.0,
        "USD": 3.73,
        "EUR": 4.28,
        "RUB": 0.047,
        "ARS": 0.0031
    }

    def generate_spending_report(self, accounts_list: list[BankAccount]) -> None:
        transactions = []
        for account in accounts_list:
            for t in account.transactions:
                if t.transaction_type == "-":
                    rate = self.EXCHANGE_RATES_TO_PLN.get(account.currency.code, 1)
                    amount_pln = t.amount * rate
                    transactions.append((t.date, amount_pln))

        if not transactions:
            raise NoReportableTransactionsError("No expense transactions to report.")

        transactions.sort(key=lambda x: x[0])
        dates = [t[0] for t in transactions]
        amounts = [t[1] for t in transactions]

        cumulative = []
        total = 0
        for amt in amounts:
            total += amt
            cumulative.append(total)

        plot.figure(figsize=(10, 5))
        plot.plot(dates, cumulative, marker='o')
        plot.xlabel("Date")
        plot.ylabel("Cumulative Spending (PLN)")
        plot.title("Cumulative Expenses Over Time")
        plot.grid(True)
        plot.tight_layout()
        plot.show()

    def generate_category_report(self, accounts_list: list[BankAccount]) -> None:
        category_totals = defaultdict(float)
        for account in accounts_list:
            rate = self.EXCHANGE_RATES_TO_PLN.get(account.currency.code, 1)
            for t in account.transactions:
                if t.transaction_type == "-":
                    category = getattr(t, "category", "Other") or "Other"
                    amount_pln = t.amount * rate
                    category_totals[str(category)] += amount_pln

        if not category_totals:
            raise NoReportableTransactionsError("No expense transactions to report.")

        labels = list(category_totals.keys())
        values = list(category_totals.values())

        plot.figure(figsize=(8, 8))
        plot.pie(values, labels=labels, autopct='%1.1f%%', startangle=140)
        plot.title("Spending by Category (PLN)")
        plot.tight_layout()
        plot.show()

    def export_spending_report_to_excel(self, accounts_list: list[BankAccount],
                                        filename=f"{REPORTS_DIR}spending_report.xlsx") -> None:
        transactions = []
        for account in accounts_list:
            for t in account.transactions:
                if t.transaction_type == "-":
                    rate = self.EXCHANGE_RATES_TO_PLN.get(account.currency.code, 1)
                    amount_pln = t.amount * rate
                    transactions.append((t.date, t.amount, account.currency.code, amount_pln, t.description,
                                         str(getattr(t, "category", "Other"))))

        if not transactions:
            raise NoReportableTransactionsError("No expense transactions to export.")

        transactions.sort(key=lambda x: x[0])

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Spending Report"

        headers = ["Date", "Original Amount", "Currency", "Amount in PLN", "Description", "Category"]
        ws.append(headers)

        for row in transactions:
            ws.append(list(row))

        for col in ws.columns:
            max_length = 0
            column = col[0].column  # Get the column name
            for cell in col:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[get_column_letter(column)].width = max_length + 2

        wb.save(filename)


    def export_category_report_to_excel(self, accounts_list: list[BankAccount],
                                        filename=f"{REPORTS_DIR}category_report.xlsx") -> None:
        category_totals = self._calculate_category_totals(accounts_list)

        if not category_totals:
            raise NoReportableTransactionsError("No expense transactions to export.")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Category Report"
        ws.append(["Category", "Total Spending (PLN)"])

        for cat, value in category_totals.items():
            ws.append([cat, value])

        self._auto_adjust_column_width(ws)
        wb.save(filename)

    def _calculate_category_totals(self, accounts_list):
        category_totals = defaultdict(float)
        for account in accounts_list:
            rate = self.EXCHANGE_RATES_TO_PLN.get(account.currency.code, 1)
            for t in account.transactions:
                if t.transaction_type == "-":
                    category = getattr(t, "category", "Other") or "Other"
                    amount_pln = t.amount * rate
                    category_totals[str(category)] += amount_pln
        return category_totals

    def _auto_adjust_column_width(self, worksheet):
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column
            for cell in col:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            worksheet.column_dimensions[get_column_letter(column)].width = max_length + 2
